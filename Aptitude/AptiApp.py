import streamlit as st
import random
import os
import base64
from openpyxl import load_workbook
from PIL import Image as PILImage
import io
import time  # For time tracking
import pymongo
from pymongo import MongoClient
from datetime import datetime
import cv2  # For face and eye detection

# Import for live camera feed
from streamlit_webrtc import webrtc_streamer, VideoTransformerBase, RTCConfiguration


# ---------------------------
# Video Transformer with Proctoring Enhancements and Smoothing (including eye-gaze tracking)
# ---------------------------
class VideoTransformer(VideoTransformerBase):
    def __init__(self):
        # Load Haar Cascade for face detection.
        self.face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
        )
        # Load Haar Cascade for eye detection.
        self.eye_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_eye.xml"
        )
        # Warning counters and timers.
        self.no_face_warning_count = 0
        self.multiple_face_warning_count = 0
        self.eye_gaze_warning_count = 0  # New counter for eye gaze violations.
        self.last_no_face_warning_time = time.time()
        self.last_multiple_warning_time = time.time()
        self.last_eye_gaze_warning_time = time.time()
        self.test_terminated = False

        # Smoothing parameters: count consecutive frames.
        self.no_face_frames = 0
        self.multiple_face_frames = 0
        self.eye_gaze_frames = 0  # New counter for consecutive eye gaze violations.
        self.frame_threshold = 5  # Only trigger warning after 5 consecutive frames.

        # Timing thresholds.
        self.warning_interval = 2  # seconds between warnings.
        self.warning_limit = 10  # number of warnings before termination.

        # New attribute to control proctoring activation.
        self.proctoring_enabled = False

        # NEW: Student id attribute to log violations for a particular student.
        self.student_id = None

    def transform(self, frame):
        # Only run proctoring if enabled.
        if not self.proctoring_enabled:
            return frame.to_ndarray(format="bgr24")

        img = frame.to_ndarray(format="bgr24")
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        current_time = time.time()
        violation_message = None

        # ---------------------------
        # Face Detection
        # ---------------------------
        faces = self.face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

        # Check for "no face" condition.
        if len(faces) == 0:
            self.no_face_frames += 1
        else:
            self.no_face_frames = 0

        if self.no_face_frames >= self.frame_threshold:
            if current_time - self.last_no_face_warning_time > self.warning_interval:
                self.no_face_warning_count += 1
                self.last_no_face_warning_time = current_time
                if self.student_id:
                    store_face_log(self.student_id, "No Face Detected!")
            violation_message = "No Face Detected!"

        # Check for "multiple faces" condition.
        if len(faces) > 1:
            self.multiple_face_frames += 1
        else:
            self.multiple_face_frames = 0

        if self.multiple_face_frames >= self.frame_threshold:
            if current_time - self.last_multiple_warning_time > self.warning_interval:
                self.multiple_face_warning_count += 1
                self.last_multiple_warning_time = current_time
                if self.student_id:
                    store_face_log(self.student_id, "Multiple Faces Detected!")
            violation_message = "Multiple Faces Detected!"

        # Draw rectangles around detected faces.
        for (x, y, w, h) in faces:
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

        # ---------------------------
        # Eye-Gaze Tracking (if exactly one face is detected)
        # ---------------------------
        if len(faces) == 1:
            (fx, fy, fw, fh) = faces[0]
            face_roi_gray = gray[fy:fy + fh, fx:fx + fw]
            eyes = self.eye_cascade.detectMultiScale(face_roi_gray, scaleFactor=1.1, minNeighbors=5)

            # Draw rectangles around detected eyes (in blue).
            for (ex, ey, ew, eh) in eyes:
                cv2.rectangle(img, (fx + ex, fy + ey), (fx + ex + ew, fy + ey + eh), (255, 0, 0), 2)

            violation_detected = False
            # If fewer than two eyes are detected, count as a potential violation.
            if len(eyes) < 2:
                violation_detected = True
            else:
                for (ex, ey, ew, eh) in eyes:
                    eye_roi = face_roi_gray[ey:ey + eh, ex:ex + ew]
                    eye_roi = cv2.equalizeHist(eye_roi)
                    _, thresholded = cv2.threshold(eye_roi, 30, 255, cv2.THRESH_BINARY_INV)
                    contours, _ = cv2.findContours(thresholded, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                    if contours:
                        max_contour = max(contours, key=cv2.contourArea)
                        M = cv2.moments(max_contour)
                        if M["m00"] != 0:
                            cx = int(M["m10"] / M["m00"])
                            if cx < ew / 3 or cx > 2 * ew / 3:
                                violation_detected = True
                    else:
                        violation_detected = True

            if violation_detected:
                self.eye_gaze_frames += 1
            else:
                self.eye_gaze_frames = 0

            if self.eye_gaze_frames >= self.frame_threshold:
                if current_time - self.last_eye_gaze_warning_time > self.warning_interval:
                    self.eye_gaze_warning_count += 1
                    self.last_eye_gaze_warning_time = current_time
                    if self.student_id:
                        store_face_log(self.student_id, "Not Looking at Screen!")
                violation_message = "Not Looking at Screen!"

        if violation_message:
            overlay = img.copy()
            cv2.rectangle(overlay, (0, 0), (img.shape[1], img.shape[0]), (0, 0, 255), -1)
            alpha = 0.4
            cv2.addWeighted(overlay, alpha, img, 1 - alpha, 0, img)
            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = img.shape[1] / 800
            thickness = max(2, int(img.shape[1] / 400))
            text_size, _ = cv2.getTextSize(violation_message, font, font_scale, thickness)
            text_x = (img.shape[1] - text_size[0]) // 2
            text_y = (img.shape[0] + text_size[1]) // 2
            cv2.putText(img, violation_message, (text_x, text_y), font, font_scale, (255, 255, 255), thickness,
                        cv2.LINE_AA)

        if (self.no_face_warning_count >= self.warning_limit) or \
                (self.multiple_face_warning_count >= self.warning_limit) or \
                (self.eye_gaze_warning_count >= self.warning_limit):
            self.test_terminated = True

        return img


# ---------------------------
# MongoDB Connection and Logging
# ---------------------------
def db_connect():
    client = pymongo.MongoClient("mongodb://localhost:27017/")
    return client['quiz_system']


def store_face_log(student_id, message):
    """Log proctoring violations in the database."""
    db = db_connect()
    collection = db["face_logs"]
    log_data = {
        "student_id": student_id,
        "timestamp": datetime.now(),
        "violation": message
    }
    try:
        collection.insert_one(log_data)
    except Exception as e:
        st.error(f"Error logging data: {e}")


# ---------------------------
# Quiz Functions and Database Operations
# ---------------------------
general_categories = [
    "aptitude",
    "data-interpretation",
    "verbal-ability",
    "logical-reasoning",
    "verbal-reasoning",
    "non-verbal-reasoning"
]

technical_categories = [
    "c-programming",
    "cpp-programming",
    "c-sharp-programming",
    "java-programming"
]


def get_test_number(username, category):
    db = db_connect()
    collection = db["apti_test"]
    latest_test_cursor = collection.find({"student_id": username, "category": category}).sort("timestamp",
                                                                                              pymongo.DESCENDING).limit(
        1)
    latest_test = list(latest_test_cursor)
    return latest_test[0]["test_no"] + 1 if latest_test else 1


def get_test_wise_accuracy(username, category, test_no):
    db = db_connect()
    collection = db['apti_test']
    test_details = collection.find({"student_id": username, "category": category, "test_no": test_no})
    correct_answers = 0
    total_questions = 0
    for test in test_details:
        correct_answers += test['marks_achieved']
        total_questions += test['no_of_questions']
    accuracy = (correct_answers / total_questions) * 100 if total_questions > 0 else 0
    return round(accuracy, 2)


def get_average_accuracy(username, category, current_accuracy=None):
    db = db_connect()
    collection = db['apti_test']
    test_details = collection.find({"student_id": username, "category": category})
    total_accuracy = 0
    test_count = 0
    for test in test_details:
        test_accuracy = get_test_wise_accuracy(username, category, test['test_no'])
        total_accuracy += test_accuracy
        test_count += 1
    if current_accuracy is not None:
        total_accuracy += current_accuracy
    avg_test_accuracy = (total_accuracy / (test_count + 1)) if test_count > 0 else current_accuracy
    return round(avg_test_accuracy, 2)


def store_test_details(username, test_no, category, no_of_questions, marks_achieved, time_taken, avg_test_accuracy):
    db = db_connect()
    collection = db['apti_test']
    existing_test = collection.find_one({
        "student_id": username,
        "test_no": test_no,
        "category": category
    })
    if existing_test:
        return
    test_data = {
        "student_id": username,
        "timestamp": datetime.now(),
        "category": category,
        "test_no": test_no,
        "no_of_questions": no_of_questions,
        "marks_achieved": marks_achieved,
        "time_taken": time_taken,
        "avg_test_accuracy": avg_test_accuracy
    }
    try:
        collection.insert_one(test_data)
        st.success("Test details stored successfully.")
    except Exception as e:
        st.error(f"Error inserting test data: {e}")


def load_questions(category):
    questions = []
    category_list = general_categories if category == 'General' else technical_categories
    for subcategory in category_list:
        file_name = f"{subcategory}.xlsx"
        file_path = os.path.join(os.path.dirname(__file__), file_name)
        if not os.path.exists(file_path):
            continue  # Skip if file doesn't exist
        wb = load_workbook(file_path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            question_no = row[0].value
            question_text = row[1].value
            options = row[2].value
            answer = row[3].value
            explanation = row[4].value
            img_path = None
            if question_text:
                for img in sheet._images:
                    if img.anchor._from.row == row[0].row - 1:
                        img_stream = io.BytesIO()
                        pil_image = PILImage.open(io.BytesIO(img._data()))
                        pil_image.save(img_stream, format='PNG')
                        img_stream.seek(0)
                        image_data = base64.b64encode(img_stream.read()).decode('utf-8')
                        img_path = f"data:image/png;base64,{image_data}"
                        break
            if question_no and question_text and options and answer:
                if subcategory == 'non-verbal-reasoning':
                    options_list = options.splitlines()
                else:
                    options_list = options.split(';')
                options_list = [option.strip() for option in options_list]
                labeled_options = {chr(65 + i): option for i, option in enumerate(options_list)}
                correct_label = answer
                for label, option in labeled_options.items():
                    if option.strip().lower() == answer.strip().lower():
                        correct_label = label
                        break
                if correct_label is None:
                    correct_label = "Unknown"
                questions.append({
                    'question_no': question_no,
                    'question_text': question_text,
                    'image_data': img_path,
                    'options': options_list,
                    'labeled_options': labeled_options,
                    'correct_answer': correct_label,
                    'explanation': explanation.strip() if explanation else "No explanation available."
                })
    return questions


def rerun_app():
    if hasattr(st, 'rerun'):
        st.rerun()
    elif hasattr(st, 'experimental_rerun'):
        st.experimental_rerun()
    else:
        st.error("Rerun not supported in this version of Streamlit. Please upgrade Streamlit.")


# ---------------------------
# Streamlit UI
# ---------------------------
st.title("🧠 AptiQuiz - Practice Your Skills!")

# RTC configuration for webrtc_streamer.
RTC_CONFIGURATION = RTCConfiguration({
    "iceServers": [{"urls": ["stun:stun.l.google.com:19302"]}]
})

# Sidebar: Live Camera Feed, Warnings, and Question Navigation
st.sidebar.title("Live Camera Feed")
with st.sidebar:
    try:
        camera = webrtc_streamer(
            key="camera",
            video_transformer_factory=VideoTransformer,
            rtc_configuration=RTC_CONFIGURATION,
            async_processing=True,  # Enable asynchronous processing for stability
            media_stream_constraints={"video": True, "audio": False}
        )
    except Exception as e:
        st.error(f"Error initializing camera: {e}")
        camera = None

    # Display warning counts for face detection and eye gaze.
    if camera and hasattr(camera, "video_transformer") and camera.video_transformer is not None:
        st.markdown(f"**No Face Warnings:** {camera.video_transformer.no_face_warning_count}")
        st.markdown(f"**Multiple Face Warnings:** {camera.video_transformer.multiple_face_warning_count}")
        st.markdown(f"**Eye-Gaze Warnings:** {camera.video_transformer.eye_gaze_warning_count}")

    st.markdown("---")

    # Custom container for question navigation with circular buttons
    st.markdown('<div id="question-nav">', unsafe_allow_html=True)
    st.markdown("""
    <style>
    #question-nav .stButton button {
        border-radius: 50% !important;
        width: 40px !important;
        height: 40px !important;
        padding: 0 !important;
        font-size: 14px !important;
        margin: 2px;
    }
    </style>
    """, unsafe_allow_html=True)

    if "questions" in st.session_state and st.session_state.questions:
        st.markdown("### Jump to Question")
        num_questions = len(st.session_state.questions)
        cols_per_row = 5
        rows = (num_questions + cols_per_row - 1) // cols_per_row
        for row in range(rows):
            cols = st.columns(cols_per_row)
            for col_index in range(cols_per_row):
                question_index = row * cols_per_row + col_index
                if question_index < num_questions:
                    button_label = str(question_index + 1)

                    # Check if the question is answered
                    is_answered = st.session_state.user_answers[question_index] is not None

                    # Create button with conditional color based on answered status
                    button_style = "background-color: #4CAF50; color: white;" if is_answered else ""

                    button_html = f"""
                    <button style="border-radius: 50%; width: 40px; height: 40px; font-size: 14px; {button_style}">
                        {button_label}
                    </button>
                    """
                    
                    # Display the button in the column
                    with cols[col_index]:
                        st.markdown(button_html, unsafe_allow_html=True)

                        # Button action - if clicked, go to the respective question
                        if st.button(button_label, key=f"qbutton_{question_index}"):
                            st.session_state.current_question = question_index
                            rerun_app()

    st.markdown('</div>', unsafe_allow_html=True)

    # Update camera_started flag based on camera state.
    # Modified: If the quiz has already started, we assume quiz section should be shown.
    if st.session_state.get("started", False) or (
            camera and hasattr(camera, "state") and getattr(camera.state, "playing", False)):
        st.session_state.camera_started = True
    else:
        st.session_state.camera_started = False

# Determine whether to show the quiz:
# Show quiz if quiz has started OR if camera is started.
if not (st.session_state.get("started", False) or st.session_state.get("camera_started", False)):
    st.warning("Please start your camera from the left sidebar before starting the quiz!")
    if st.button("Retry Camera"):
        rerun_app()
else:
    with st.container():
        if "started" not in st.session_state or not st.session_state.started:
            username = st.text_input("Enter your username:")
            category = st.radio("Choose a category:", ['General', 'Technical'])
            if username:
                test_no = get_test_number(username, category)
                st.write(f"Test Number: {test_no}")
            test_type = st.radio("Select Your Test Mode:",
                                 ["⚡ Quick Challenge (10 Questions)", "🏆 Full Test (30 Questions)"])
            no_of_questions = 10 if "Quick Challenge" in test_type else 30
            if st.button("Start Quiz"):
                st.session_state.started = True
                st.session_state.username = username
                st.session_state.category = category
                st.session_state.test_no = test_no
                st.session_state.no_of_questions = no_of_questions
                st.session_state.questions = load_questions(category)
                random.shuffle(st.session_state.questions)
                st.session_state.questions = st.session_state.questions[:no_of_questions]
                st.session_state.current_question = 0
                st.session_state.user_answers = [None] * st.session_state.no_of_questions
                st.session_state.start_time = time.time()
                st.session_state.test_submitted = False
                if camera is not None and hasattr(camera, "video_transformer") and camera.video_transformer is not None:
                    camera.video_transformer.proctoring_enabled = True
                    # NEW: Set the student_id in the video transformer for logging purposes.
                    camera.video_transformer.student_id = st.session_state.username

    st.markdown("---")

    if ("questions" in st.session_state and
            (st.session_state.get("test_submitted", False) or
             st.session_state.current_question == len(st.session_state.questions) or
             st.session_state.get("test_terminated", False))):
        # Disable proctoring and reset warnings once the test is over.
        if camera and hasattr(camera, "video_transformer"):
            camera.video_transformer.proctoring_enabled = False
            camera.video_transformer.no_face_warning_count = 0
            camera.video_transformer.multiple_face_warning_count = 0
            camera.video_transformer.eye_gaze_warning_count = 0

        with st.container():
            st.markdown("---")
            end_time = time.time()
            time_taken = round(end_time - st.session_state.start_time, 2)
            score = 0
            for i, q in enumerate(st.session_state.questions):
                if st.session_state.user_answers[i] == q["correct_answer"]:
                    score += 1
            st.header("🎉 Quiz Completed!")
            st.write(f"**Your Score:** {score} / {len(st.session_state.questions)}")
            st.write(f"**Time Taken:** {time_taken} seconds")
            if st.session_state.get("test_terminated", False):
                st.error("Your test was terminated due to proctoring violations.")
            st.subheader("Correct Answers and Explanations:")
            for i, q in enumerate(st.session_state.questions):
                if i >= len(st.session_state.user_answers):
                    break
                user_ans = st.session_state.user_answers[i]
                correct_ans = q["correct_answer"]
                st.markdown(f"**Q{i + 1}:** {q['question_text']}")
                if q["image_data"]:
                    st.image(q["image_data"], use_container_width=True)
                st.markdown(f"**✅ Correct Answer:** {correct_ans}")
                st.markdown(f"**❌ Your Answer:** {user_ans}")
                st.markdown(f"**💡 Explanation:** {q['explanation']}")
                st.write("---")
            current_accuracy = get_test_wise_accuracy(
                st.session_state.username, st.session_state.category, st.session_state.test_no
            )
            avg_test_accuracy = get_average_accuracy(
                st.session_state.username, st.session_state.category, current_accuracy
            )
            store_test_details(
                st.session_state.username,
                st.session_state.test_no,
                st.session_state.category,
                st.session_state.no_of_questions,
                score,
                time_taken,
                avg_test_accuracy
            )
            if st.button("Try Again"):
                st.session_state.started = False
                for key in ["username", "category", "test_no", "questions", "current_question", "user_answers",
                            "start_time", "test_terminated", "test_submitted"]:
                    if key in st.session_state:
                        del st.session_state[key]
                rerun_app()

    elif st.session_state.get("started", False):
        # If the quiz is in progress, display the current question.
        with st.container():
            current_index = st.session_state.current_question
            question_data = st.session_state.questions[current_index]
            with st.form(key="question_form"):
                st.header(f"Question {current_index + 1} / {len(st.session_state.questions)}")
                st.write(question_data["question_text"])
                if question_data["image_data"]:
                    st.image(question_data["image_data"], use_container_width=True)
                options = list(question_data["labeled_options"].keys())
                default_answer = st.session_state.user_answers[current_index]
                default_index = options.index(default_answer) if default_answer in options else 0
                user_choice = st.radio(
                    "Choose your answer:",
                    options,
                    index=default_index,
                    format_func=lambda x: f"{x}: {question_data['labeled_options'][x]}"
                )
                col1, col2, col3 = st.columns(3)
                with col1:
                    prev_pressed = st.form_submit_button("Previous")
                with col2:
                    next_pressed = st.form_submit_button("Next")
                with col3:
                    submit_pressed = st.form_submit_button("Submit Test")
                st.session_state.user_answers[current_index] = user_choice
                if prev_pressed:
                    if current_index > 0:
                        st.session_state.current_question -= 1
                    else:
                        st.warning("This is the first question.")
                    rerun_app()
                elif next_pressed:
                    if current_index < len(st.session_state.questions) - 1:
                        st.session_state.current_question += 1
                    else:
                        st.warning("This is the last question.")
                    rerun_app()
                elif submit_pressed:
                    st.session_state.test_submitted = True
                    rerun_app()
